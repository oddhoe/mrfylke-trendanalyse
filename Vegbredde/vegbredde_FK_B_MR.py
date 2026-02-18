"""
Henter VT 838 (Vegbredde, beregnet) i Møre og Romsdal
med Dekkebredde < 6.0 m OG overlapp med Funksjonsklasse B (VT 912).
Eksporterer til GeoPackage (.gpkg) og Excel (.xlsx) for ArcGIS Pro.

Egenskaps-ID-er:
  9537  = Dekkebredde (VT 838)
  11216 = Funksjonsklasse (VT 912)
  18511=A, 18512=B, 18513=C, 18514=D, 18515=E
"""

import requests
import geopandas as gpd
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from shapely import wkt
from shapely.geometry import LineString, MultiLineString, Point
from datetime import datetime

BASE_URL = "https://nvdbapiles.atlas.vegvesen.no"
HEADERS  = {"X-Client": "nvdb-bredde-FK-B-gpkg", "Accept": "application/json"}


# ── Hjelpefunksjonar ──────────────────────────────────────────────────────────

def hent_fylkesnummer(navn):
    r = requests.get(f"{BASE_URL}/omrader/api/v4/fylker", headers=HEADERS)
    r.raise_for_status()
    for f in r.json():
        if navn.lower() in f["navn"].lower():
            print(f"  Fylke: {f['navn']} (nr={f['nummer']})")
            return f["nummer"]
    raise ValueError(f"Fant ikke: {navn}")


def hent_vegobjekter(vt_id, fylke_nr, fk_enum):
    start_url = (
        f"{BASE_URL}/vegobjekter/{vt_id}"
        f"?fylke={fylke_nr}"
        f'&egenskap="egenskap(9537)<6.0"'
        f"&overlapp=912(egenskap(11216)={fk_enum})"
        f"&inkluder=egenskaper,lokasjon,vegsegmenter,geometri"
        f"&inkludergeometri=utledet"
        f"&srid=4326"
        f"&antall=1000"
    )

    objekter = []
    url = start_url
    side = 1

    while url:
        r = requests.get(url, headers=HEADERS)
        if not r.ok:
            print(f"  FEIL {r.status_code}: {r.text[:300]}")
            r.raise_for_status()
        data = r.json()
        batch = data.get("objekter", [])
        objekter.extend(batch)
        print(f"  Side {side}: {len(batch)} obj, totalt {len(objekter)}")
        side += 1
        neste = data.get("metadata", {}).get("neste")
        if not neste or not batch:
            break
        url = neste.get("href")

    print(f"\n  Totalt: {len(objekter)} objekter")
    return objekter


EG_MAP = {
    9537:  "dekkebredde_m",
    9538:  "dekkebredde_min",
    9536:  "dekkebredde_maks",
    10248: "dekkebredde_median",
    10249: "dekkebredde_normal",
    9797:  "vegbredde_m",
    9800:  "kjorebane_m",
    9534:  "grunnlag",
    9535:  "datafangst_fkb",
    9533:  "dato_grunnlag_eldste",
    13098: "dato_grunnlag_nyeste",
}

EXCEL_KOLONNER = {
    "nvdb_id":              "NVDB objekt-ID",
    "funksjonsklasse":      "Funksjonsklasse",
    "fra_dato":             "Fra dato",
    "dekkebredde_m":        "Dekkebredde (m)",
    "dekkebredde_min":      "Dekkebredde, min (m)",
    "dekkebredde_maks":     "Dekkebredde, maks (m)",
    "dekkebredde_median":   "Dekkebredde, median (m)",
    "dekkebredde_normal":   "Dekkebredde, normal (m)",
    "vegbredde_m":          "Vegbredde (m)",
    "kjorebane_m":          "Kjørebanebredde (m)",
    "grunnlag":             "Grunnlag",
    "datafangst_fkb":       "Datafangstmetode FKB",
    "dato_grunnlag_eldste": "Dato grunnlagsdata, eldste",
    "dato_grunnlag_nyeste": "Dato grunnlagsdata, nyeste",
    "fylke":                "Fylke",
    "kommune":              "Kommune",
    "lengde_m":             "Lengde stedfesting (m)",
    "vegreferanse":         "Vegreferanse",
    "vegnummer":            "Vegnummer",
    "vegkategori":          "Vegkategori",
}


def parse_vegsegmenter(segs):
    kortformer, vegnumre, kategorier = [], [], []
    for s in segs:
        vsr = s.get("vegsystemreferanse", {})
        kf  = vsr.get("kortform", "")
        if kf and kf not in kortformer:
            kortformer.append(kf)
        vs  = vsr.get("vegsystem", {})
        nr  = vs.get("nummer")
        if nr is not None and str(nr) not in vegnumre:
            vegnumre.append(str(nr))
        kat = vs.get("vegkategori", "")
        if kat and kat not in kategorier:
            kategorier.append(kat)
    return "; ".join(kortformer), "; ".join(vegnumre), "; ".join(kategorier)


def parse_objekt(obj, funksjonsklasse_label):
    rad = {
        "nvdb_id":         obj.get("id"),
        "funksjonsklasse": funksjonsklasse_label,
        "fra_dato":        obj.get("metadata", {}).get("startdato"),
    }

    for eg in obj.get("egenskaper", []):
        felt = EG_MAP.get(eg.get("id"))
        if felt:
            rad[felt] = eg.get("verdi")

    lok = obj.get("lokasjon", {})
    rad["fylke"]    = lok.get("fylker",   [None])[0]
    rad["kommune"]  = lok.get("kommuner", [None])[0]
    rad["lengde_m"] = lok.get("lengde")

    segs = obj.get("vegsegmenter", [])
    rad["vegreferanse"], rad["vegnummer"], rad["vegkategori"] = parse_vegsegmenter(segs)

    geom_obj     = obj.get("geometri") or {}
    wkt_kandidat = geom_obj.get("wkt") or geom_obj.get("geometri")
    if not wkt_kandidat:
        for eg in obj.get("egenskaper", []):
            if eg.get("datatype") in ("GeomPunkt", "GeomLinje", "GeomFlate"):
                wkt_kandidat = eg.get("verdi")
                break

    rad["_wkt"] = wkt_kandidat
    return rad


def byt_xy(geom):
    """Byter om x og y i geometrien, bevarar evt. z."""
    if geom is None:
        return None

    def snukoord(coords):
        if len(coords[0]) == 3:
            return [(y, x, z) for x, y, z in coords]
        return [(y, x) for x, y in coords]

    if geom.geom_type == "LineString":
        return LineString(snukoord(list(geom.coords)))
    elif geom.geom_type == "MultiLineString":
        return MultiLineString([snukoord(list(line.coords)) for line in geom.geoms])
    elif geom.geom_type == "Point":
        c = list(geom.coords)[0]
        return Point(c[1], c[0]) if len(c) == 2 else Point(c[1], c[0], c[2])
    else:
        print(f"  Advarsel: ukjent geometritype {geom.geom_type} – ikkje bytta")
        return geom


def sjekk_koordinatar(gdf, n=3):
    print("\n--- Koordinat-sjekk ---")
    feil_teller = 0
    ok_teller   = 0

    for _, row in gdf[gdf.geometry.notna()].head(n).iterrows():
        geom = row.geometry
        try:
            coords = list(geom.coords) if hasattr(geom, "coords") else list(geom.geoms[0].coords)
        except Exception:
            print(f"  [{row['nvdb_id']}] Kunne ikkje lese koordinatar")
            continue

        first = coords[0]
        x, y  = first[0], first[1]
        vref  = row.get("vegreferanse", str(row["nvdb_id"]))

        print(f"  [{row['nvdb_id']}] Rå koordinatar: {first}")

        if 5 <= x <= 32 and 57 <= y <= 72:
            print(f"  ✓ {vref}  →  lon={x:.5f}, lat={y:.5f}  (korrekt)")
            ok_teller += 1
        elif 57 <= x <= 72 and 5 <= y <= 32:
            print(f"  ✗ {vref}  →  x={x:.5f}, y={y:.5f}  (LAT/LON BYTTET OM – fiksar)")
            feil_teller += 1
        else:
            print(f"  ? {vref}  →  x={x:.5f}, y={y:.5f}  (ukjent område)")

    koordinatar_ok = feil_teller == 0
    print(f"  Resultat: {ok_teller} ok, {feil_teller} feil av {n} testa\n")
    return koordinatar_ok


def bygg_geodataframe(rader):
    df = pd.DataFrame(rader)
    print(f"  Objekter med geometri:  {df['_wkt'].notna().sum()}")
    print(f"  Objekter UTEN geometri: {df['_wkt'].isna().sum()}")
    df["geometry"] = df["_wkt"].apply(
        lambda w: wkt.loads(w) if isinstance(w, str) else None
    )
    df = df.drop(columns=["_wkt"])
    gdf = gpd.GeoDataFrame(df, geometry="geometry", crs="EPSG:4326")
    return gdf


def hent_geometri_enkeltvis(nvdb_ids):
    resultat = {}
    print(f"  Henter geometri enkeltvis for {len(nvdb_ids)} objekter...")
    for nvdb_id in nvdb_ids:
        url = (
            f"{BASE_URL}/vegobjekter/838/{nvdb_id}"
            f"?inkluder=geometri&inkludergeometri=utledet&srid=4326"
        )
        r = requests.get(url, headers=HEADERS)
        if r.ok:
            data     = r.json()
            geom_obj = data.get("geometri") or {}
            wkt_str  = geom_obj.get("wkt") or geom_obj.get("geometri")
            if wkt_str:
                resultat[nvdb_id] = wkt_str
    print(f"  Fikk geometri for {len(resultat)} av {len(nvdb_ids)}")
    return resultat


def skriv_xlsx(gdf, filnavn, funksjonsklasse):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Vegbredde_FK_{funksjonsklasse}"

    kolonner_intern = [k for k in EXCEL_KOLONNER.keys() if k in gdf.columns]
    kolonner_excel  = [EXCEL_KOLONNER[k] for k in kolonner_intern]
    ws.append(kolonner_excel)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="BFBFBF")

    for _, row in gdf.iterrows():
        ws.append([row.get(k) for k in kolonner_intern])

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

    ws.freeze_panes = "A2"
    wb.save(filnavn)
    print(f"  Lagret: {filnavn}")


# ── Hovudprogram ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    FUNKSJONSKLASSE = "B"
    FK_ENUM         = 18512      # 18511=A, 18512=B, 18513=C, 18514=D, 18515=E

    print("Finner fylkesnummer...")
    fylke_nr = hent_fylkesnummer("Møre og Romsdal")

    print(f"Henter vegbredde < 6.0 m med FK={FUNKSJONSKLASSE}...")
    objekter = hent_vegobjekter(838, fylke_nr, FK_ENUM)

    print("Parser...")
    rader = [parse_objekt(o, FUNKSJONSKLASSE) for o in objekter]

    print("Bygger GeoDataFrame...")
    gdf = bygg_geodataframe(rader)

    # Koordinat-sjekk og automatisk fiks
    koordinatar_ok = sjekk_koordinatar(gdf, n=3)
    if not koordinatar_ok:
        print("  Fiksar koordinatrekkefølge (byter x/y)...")
        gdf["geometry"] = gdf["geometry"].apply(byt_xy)
        print("  Verifiserer etter fiks:")
        sjekk_koordinatar(gdf, n=3)

    # Fallback: hent geometri enkeltvis for objekt som manglar
    mangler_mask = gdf["geometry"].isna()
    if mangler_mask.sum() > 0:
        ids_uten = gdf.loc[mangler_mask, "nvdb_id"].tolist()
        ekstra   = hent_geometri_enkeltvis(ids_uten)
        for nvdb_id, wkt_str in ekstra.items():
            idx = gdf.index[gdf["nvdb_id"] == nvdb_id]
            gdf.loc[idx, "geometry"] = wkt.loads(wkt_str)

    gdf_ok = gdf[gdf["geometry"].notna()].copy()
    if len(gdf_ok) < len(gdf):
        print(f"  {len(gdf) - len(gdf_ok)} rader droppes frå .gpkg (ingen geometri)")

    ts = datetime.now().strftime("%Y%m%d_%H%M")

    gpkg_fil = f"vegbredde_FK_{FUNKSJONSKLASSE}_MR_{ts}.gpkg"
    print(f"Skriver {gpkg_fil} ({len(gdf_ok)} features)...")
    gdf_ok.to_file(gpkg_fil, driver="GPKG", layer=f"vegbredde_FK_{FUNKSJONSKLASSE}")
    print(f"  Lagret: {gpkg_fil}")

    xlsx_fil = f"vegbredde_FK_{FUNKSJONSKLASSE}_MR_{ts}.xlsx"
    print(f"Skriver {xlsx_fil} ({len(gdf)} rader)...")
    skriv_xlsx(gdf, xlsx_fil, FUNKSJONSKLASSE)

    print("Ferdig!")
